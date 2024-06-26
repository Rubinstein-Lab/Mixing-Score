# This code has been written by Gil Wiseglass and Nadir Boni from the Rotem Rubinstein lab, Tel Aviv University, Israel.
# This code has been used for the data presented in the paper "clustered protocadherin cis-interactions are
# required for combinatorial cell-cell recognition underlying neuronal self-avoidance".

import cv2
import openpyxl
from PIL import Image


# Function to parse image into squares of a given size
def parse_image(img, square_size):
    # Get dimensions of image
    width, height = img.size

    # Calculate number of squares that fit horizontally and vertically
    num_horizontal = width // square_size
    num_vertical = height // square_size

    # Initialize list to hold square images
    squares = []

    # Iterate over squares and add to list
    for i in range(num_vertical):
        for j in range(num_horizontal):
            # Calculate coordinates of top-left corner of square
            x = j * square_size
            y = i * square_size

            # Crop square image from original image
            square = img.crop((x, y, x + square_size, y + square_size))

            # Add square image to list
            squares.append(square)

    return squares


# Function to calculate number of squares with white pixels in both images
def calculate_num_both_white(img1_path, img2_path, square_size):
    # Open images
    img1 = Image.open(img1_path)
    img2 = Image.open(img2_path)

    # Parse images into squares
    img1_squares = parse_image(img1, square_size)
    img2_squares = parse_image(img2, square_size)

    # Initialize counters for squares
    num_both_white = 0  # number of white squares in both images (1 intersection 2)
    num_white_img1 = 0  # number of white squares in 1 (1)
    num_white_img2 = 0  # number of white squares in 2 (2)
    num_white = 0  # total number of white squares (1 union 2)

    # Iterate over squares and check for white pixels in both images
    for img1_square, img2_square in zip(img1_squares, img2_squares):
        img1_flag = False
        img2_flag = False

        # Iterate over pixels and check if they're white in either image
        for pixel1 in img1_square.getdata():
            if pixel1 == 255:
                img1_flag = True
                break

        for pixel2 in img2_square.getdata():
            if pixel2 == 255:
                img2_flag = True
                break

        if img1_flag and img2_flag:
            num_both_white += 1
            num_white_img1 += 1
            num_white_img2 += 1
            num_white += 1

        elif img1_flag:
            num_white_img1 += 1
            num_white += 1

        elif img2_flag:
            num_white_img2 += 1
            num_white += 1

    # Return number of squares with white pixels in both images, total number of squares parsed
    return num_both_white, num_white_img1, num_white_img2, num_white


square_size = 64

# Open the input and output workbooks
input_workbook = openpyxl.load_workbook("[InputPath].xlsx")  # The Input directory
output_workbook = openpyxl.Workbook()
output_worksheet = output_workbook.active

# Iterate over the rows in the input worksheet
for row in input_workbook.active.iter_rows(min_row=1, max_col=2, values_only=True):
    # Get the image locations from the current row
    image1_location = row[0]
    image2_location = row[1]

    # Load the images
    image1 = cv2.imread(image1_location, cv2.IMREAD_GRAYSCALE)
    image2 = cv2.imread(image2_location, cv2.IMREAD_GRAYSCALE)

    # Check that the images have the same dimensions
    assert image1.shape == image2.shape

    num_both_white, num_white_img1, num_white_img2, num_white = calculate_num_both_white(image1_location, image2_location, square_size)

    # Calculating the mixing index score
    mixing_score = num_both_white/num_white

    # Write the input and output image locations and the number of squares with white pixels in both images and total
    # squares to the output worksheet
    output_worksheet.append(row + (num_both_white, num_white_img1, num_white_img2, num_white, mixing_score))

    # Save the output workbook
    output_workbook.save("[OutputPath].xlsx")  # The output directory

